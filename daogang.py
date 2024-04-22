import openpyxl


def daogang(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Iterate over rows
    for row_num in range(2, ws.max_row + 1):
        # Get values from columns H and M
        i_value = ws.cell(row=row_num, column=8).value
        p_value = ws.cell(row=row_num, column=13).value

        # Check if values are the same
        if i_value == p_value:
            # Delete the row
            ws.delete_rows(row_num, 1)

    # Save the changes to the original file
    wb.save(path+"_update.xlsx")