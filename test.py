import openpyxl
import os

from selenium import webdriver

import daogang
import myselenium

# Load the Excel file


file_path = '/Users/adiosrefrain/PycharmProjects/excel/铁矿_全球海漂_分品种_20240422.xlsx'

if "到港" in os.path.basename(file_path):
    daogang.daogang(file_path)
if "海漂" in os.path.basename(file_path):
    chrome_options = webdriver.ChromeOptions()
    # chrome_options.add_argument('--headless')
    driver = webdriver.Chrome(options=chrome_options)
    driver.get("https://www.shipxy.com/")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Create a new Excel file to save the modified rows
    # 创建新的工作簿作为模板的副本
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    delete_wb = openpyxl.Workbook()
    delete_ws = new_wb.active

    # Iterate over rows
    for row_num in range(2, ws.max_row + 1):
        # Get values from columns H and M
        i_value = ws.cell(row=row_num, column=9).value
        p_value = ws.cell(row=row_num, column=16).value

        # Check if values are the same
        if i_value == p_value:

            delete_row_values = [cell.value for cell in ws[row_num]]
            delete_ws.append(delete_row_values)
            delete_ws.cell(row=row_num, column=26).value = '到货国和发货国相同'

            ws.delete_rows(row_num, 1)
        elif p_value == '中国':
            delete_row_values = [cell.value for cell in ws[row_num]]
            delete_ws.append(delete_row_values)
            delete_ws.cell(row=row_num, column=26).value = '发货国是中国'

            ws.delete_rows(row_num, 1)

        else:
            a_value = ws.cell(row=row_num, column=1).value
            myselenium.get_country(driver, a_value,i_value, ws, row_num)

    wb.save(file_path + "_update.xlsx")
    new_wb.save(file_path + "_modified.xlsx")
    delete_wb.save(file_path + "_delete.xlsx")
    driver.quit()
